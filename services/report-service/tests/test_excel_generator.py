"""Tests for Excel generator."""

import os
import tempfile

from openpyxl import load_workbook

from app.generators.excel_generator import generate_excel


class TestExcelGenerator:
    def test_generate_single_sheet(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "test.xlsx")
            sheets_data = [{
                "name": "Ventes",
                "headers": ["Référence", "Client", "Montant"],
                "rows": [
                    ["VTE-001", "Client A", "100 000 XOF"],
                    ["VTE-002", "Client B", "250 000 XOF"],
                ],
            }]
            result = generate_excel("Test Report", sheets_data, output_path)
            assert os.path.exists(result)

            wb = load_workbook(result)
            assert "Ventes" in wb.sheetnames
            ws = wb["Ventes"]
            # Row 1: company name, Row 2: subtitle, Row 3: blank, Row 4: headers
            assert ws.cell(row=4, column=1).value == "Référence"
            assert ws.cell(row=5, column=2).value == "Client A"

    def test_generate_with_summary(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "summary.xlsx")
            sheets_data = [{
                "name": "Data",
                "headers": ["A", "B"],
                "rows": [["1", "2"]],
            }]
            summary = {"Total": "100 000 XOF", "Count": "5"}
            result = generate_excel("Résumé", sheets_data, output_path, summary=summary)
            assert os.path.exists(result)

            wb = load_workbook(result)
            assert "Résumé" in wb.sheetnames  # summary sheet
            assert "Data" in wb.sheetnames

    def test_generate_multiple_sheets(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "multi.xlsx")
            sheets_data = [
                {"name": "Ventes", "headers": ["Ref"], "rows": [["V1"]]},
                {"name": "Dépenses", "headers": ["Ref"], "rows": [["D1"]]},
            ]
            result = generate_excel("Multi", sheets_data, output_path)
            wb = load_workbook(result)
            assert "Ventes" in wb.sheetnames
            assert "Dépenses" in wb.sheetnames

    def test_generate_with_totals(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "totals.xlsx")
            sheets_data = [{
                "name": "Test",
                "headers": ["Item", "Montant"],
                "rows": [["A", "100"], ["B", "200"]],
                "totals": ["TOTAL", "300"],
            }]
            result = generate_excel("Totals", sheets_data, output_path)
            wb = load_workbook(result)
            ws = wb["Test"]
            # Data starts at row 5 (after title, subtitle, blank, headers)
            # 2 data rows (5, 6), totals at row 7
            assert ws.cell(row=7, column=1).value == "TOTAL"
            assert ws.cell(row=7, column=2).value == "300"

    def test_empty_data(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "empty.xlsx")
            sheets_data = [{
                "name": "Empty",
                "headers": ["A", "B"],
                "rows": [],
            }]
            result = generate_excel("Empty", sheets_data, output_path)
            assert os.path.exists(result)
