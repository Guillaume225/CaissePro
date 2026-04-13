"""Excel generation using openpyxl."""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import get_settings

settings = get_settings()

HEADER_FILL = PatternFill(start_color="062A5A", end_color="062A5A", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="062A5A")
SUBTITLE_FONT = Font(name="Calibri", size=10, color="4884BD")
DATA_FONT = Font(name="Calibri", size=10)
TOTAL_FILL = PatternFill(start_color="062A5A", end_color="062A5A", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
ALT_FILL = PatternFill(start_color="EDF3F9", end_color="EDF3F9", fill_type="solid")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="4884BD"),
)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def generate_excel(
    title: str,
    sheets_data: list[dict],
    output_path: str,
    summary: dict | None = None,
) -> str:
    """
    Generate professional Excel workbook.

    sheets_data: list of {
        "name": "Sheet name",
        "headers": ["Col1", "Col2", ...],
        "rows": [[val1, val2, ...], ...],
        "totals": [val1, val2, ...] | None,
        "col_widths": [15, 20, ...] | None,
    }
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = Workbook()

    # Remove default sheet
    if wb.active:
        wb.remove(wb.active)

    for sheet_info in sheets_data:
        ws = wb.create_sheet(title=sheet_info["name"][:31])
        headers = sheet_info["headers"]
        rows = sheet_info["rows"]
        totals = sheet_info.get("totals")
        col_widths = sheet_info.get("col_widths")

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=settings.company_name)
        title_cell.font = TITLE_FONT

        # Subtitle row
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        sub_cell = ws.cell(
            row=2, column=1,
            value=f"{title} — Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
        )
        sub_cell.font = SUBTITLE_FONT

        # Blank row
        start_row = 4

        # Headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER

        # Data rows
        for row_idx, row_data in enumerate(rows, start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = LEFT
                if row_idx % 2 == 0:
                    cell.fill = ALT_FILL

        # Totals row
        if totals:
            total_row = start_row + 1 + len(rows)
            for col_idx, value in enumerate(totals, 1):
                cell = ws.cell(row=total_row, column=col_idx, value=value)
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL

        # Column widths
        if col_widths:
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
        else:
            for i in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(i)].width = 18

        # Freeze header row
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # Summary sheet
    if summary:
        ws = wb.create_sheet(title="Résumé", index=0)
        ws.merge_cells("A1:D1")
        ws.cell(row=1, column=1, value=settings.company_name).font = TITLE_FONT
        ws.merge_cells("A2:D2")
        ws.cell(row=2, column=1, value=title).font = SUBTITLE_FONT

        row = 4
        for key, value in summary.items():
            ws.cell(row=row, column=1, value=key).font = Font(name="Calibri", size=10, bold=True)
            ws.cell(row=row, column=2, value=value).font = DATA_FONT
            row += 1

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25

    wb.save(output_path)
    return output_path
